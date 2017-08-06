import requests
from bs4 import BeautifulSoup
import re
import openpyxl

temp = []

url = "http://www.euro.com.pl/telewizory-led-lcd-plazmowe,strona-1.bhtml"
source_code = requests.get(url)
plain_text = source_code.text
soup = BeautifulSoup(plain_text, "html.parser")

'''for item in soup.find_all(re.compile(r'^(h2|span)$'), {'class': re.compile(r'^(product-name|attribute-value)$')}):
    temp.extend(item.strings)

temp2 = [x.encode('utf-8') for x in temp]

with open('test params.txt', 'w') as t:
    for i in temp2:
        t.write(i)

t.close()'''

wb = openpyxl.Workbook()
ws = wb.active

tagiterator = soup.h2

row, col = 1, 1
ws.cell(row=row, column=col, value=tagiterator.getText())
tagiterator = tagiterator.find_next()

while tagiterator.find_next():
    if tagiterator.name == 'h2':
        row += 1
        col = 1
        ws.cell(row=row, column=col, value=tagiterator.getText(strip=True))
    elif tagiterator.name == 'span':
        col += 1
        ws.cell(row=row, column=col, value=tagiterator.getText(strip=True))
    tagiterator = tagiterator.find_next()

wb.save('DG3test.xlsx')

