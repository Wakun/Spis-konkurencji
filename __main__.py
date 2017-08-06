# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import time
# import xlwt
import openpyxl
import numpy as np
from url_dict import RTV_url_dict, AGD_url_dict, AGD_zabudowa_url_dict, AGD_male_url_dict, Komputery_url_dict, \
    Gry_i_konsole_url_dict, Foto_i_kamery_url_dict, Telefony_i_GPS_url_dict   # słowniki do adresów stron

dict_games = {"gry-pc": "PC", "gry-playstation-4": "PS4", "gry-xbox-one": "X1", "gry-xbox-360": "XBOX 360",
              "gry-playstation-3": "PS3"}   # słownik do rozróżniania gier

prod_temp = []  # tymczasowa lista produktów
ceny_temp = []  # tymczasowa lista cen
param_temp = []  # tymczasowa lista parametrów

# utworzenie dictionary z pliku .csv
'''lines = np.genfromtxt("csvtest.csv", delimiter=";", dtype=None)
compdict = dict()
for i in range(len(lines)):
    compdict[lines[i][0]] = lines[i][1]'''


def nazwy_produktow(max_pages, cat):
    # pobieranie nazw produktów
    page = 1

    while page <= max_pages:
        url = "http://www.euro.com.pl/" + cat + ",strona-" + str(page) + ".bhtml"
        source_code = requests.get(url)
        plain_text = source_code.text
        soup = BeautifulSoup(plain_text, "html.parser")
        for prod in soup.find_all('h2', {'class': 'product-name'}):
            if cat in dict_games:
                for p in prod.stripped_strings:
                    prod_temp.append(dict_games[cat] + " " + p)
            else:
                for p in prod.stripped_strings:
                    prod_temp.append(p)
        page += 1


def ceny_produktow(max_pages, cat, sync):
    # pobieranie cen produktów
    page = 1

    while page <= max_pages:
        url = "http://www.euro.com.pl/" + cat + ",strona-" + str(page) + ".bhtml"
        source_code = requests.get(url)
        plain_text = source_code.text
        soup = BeautifulSoup(plain_text, "html.parser")
        for cena in soup.find_all('div', {'class': "price-normal selenium-price-normal"}):
            ceny_temp.extend(cena.stripped_strings)
        for i in range(sync):
            ceny_temp.pop(-1)
        page += 1


def parametry_produktow(max_pages, cat):
    # pobieranie parametrów produktów wraz z nazwami
    page = 1

    while page <= max_pages:
        url = "http://www.euro.com.pl/" + cat + ",strona-" + str(page) + ".bhtml"
        source_code = requests.get(url)
        plain_text = source_code.text
        soup = BeautifulSoup(plain_text, "html.parser")
        for param in soup.find_all('span', {'class': "attribute-value"}):
                param_temp.extend(param.stripped_strings)
        page += 1


def make_param_temp(nazwy):
    # utworzenie listy nazw produktów z parametrami
    params = [prod.encode('utf-8') for prod in nazwy]
    return dict.fromkeys(params, [])


def get_max_pages(x):
    # ustalenie ilości stron dla danej kategorii
    max_p = []

    url = "http://www.euro.com.pl/" + x + ",strona-1.bhtml"
    source_code = requests.get(url)
    plain_text = source_code.text
    soup = BeautifulSoup(plain_text, "html.parser")
    for page in soup.find_all('a', {'class': 'paging-number'}):
        max_p.extend(page.stripped_strings)

    if max_p == []:
        return 1
    else:
        p = int(max_p[-1])
        return p


def pricesynchro(cat):
    # synchronizacja cen dla spisu konkurencji

    sync_lst = []
    num = 0

    url = "http://www.euro.com.pl/" + cat + ",strona-1.bhtml"
    source_code = requests.get(url)
    plain_text = source_code.text
    soup = BeautifulSoup(plain_text, "html.parser")
    for price in soup.find_all('div', {'class': "price-normal selenium-price-normal"}):
        sync_lst.extend(price)

    if len(sync_lst) == 34:
        num = 4
    elif len(sync_lst) == 33:
        num = 3
    elif len(sync_lst) == 30:
        return num

    return num


def param_synchro(cat):
    # synchronizacja parametrów !!nie działa

    sync_lst = []
    ilosc_prod = 0

    url = "http://www.euro.com.pl/" + cat + ",strona-1.bhtml"
    source_code = requests.get(url)
    plain_text = source_code.text
    soup = BeautifulSoup(plain_text, "html.parser")
    for param in soup.find_all('span', {'class': "attribute-value"}):
        sync_lst.extend(param.stripped_strings)

    # for prod in soup.find_all('h2', {'class': 'product-name'}):
        # ilosc_prod += 1

    return len(sync_lst) / ilosc_prod


def kategoria():
    # pobranie kategorii do spisu konkurencji z url_dict

    cat = input("Wpisz kategorię do spisania: ")

    if cat in RTV_url_dict:
        cat = RTV_url_dict[cat]
    elif cat in AGD_url_dict:
        cat = AGD_url_dict[cat]
    elif cat in AGD_zabudowa_url_dict:
        cat = AGD_zabudowa_url_dict[cat]
    elif cat in AGD_male_url_dict:
        cat = AGD_male_url_dict[cat]
    elif cat in Komputery_url_dict:
        cat = Komputery_url_dict[cat]
    elif cat in Gry_i_konsole_url_dict:
        cat = Gry_i_konsole_url_dict[cat]
    elif cat in Foto_i_kamery_url_dict:
        cat = Foto_i_kamery_url_dict[cat]
    elif cat in Telefony_i_GPS_url_dict:
        cat = Telefony_i_GPS_url_dict[cat]
    return cat


def printcat():
    # wyświetlenie wszystkich możliwych kategorii do spisania

    print("Kategorie: ")
    print("\033[93m" + "\tRTV: " + "\033[0m")
    for cat in RTV_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tAGD: " + "\033[0m")
    for cat in AGD_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tAGD do zabudowy: " + "\033[0m")
    for cat in AGD_zabudowa_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tAGD Małe: " + "\033[0m")
    for cat in AGD_male_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tKomputery: " + "\033[0m")
    for cat in Komputery_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tGry i konsole: " + "\033[0m")
    for cat in Gry_i_konsole_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tFoto i kamery: " + "\033[0m")
    for cat in Foto_i_kamery_url_dict:
        print("\t\t-" + cat)
    print("\033[93m" + "\tTelefony i GPS: " + "\033[0m")
    for cat in Telefony_i_GPS_url_dict:
        print("\t\t-" + cat)


def saveresults_spis(pages, cat, sync):
    # zapisanie wyników spisu konkurencji do pliku excel

    nazwy_produktow(pages, cat)
    ceny_produktow(pages, cat, sync)

    lista_prod = [x.encode('utf-8') for x in prod_temp]
    lista_cen = [x.encode('utf-8') for x in ceny_temp]

    # zapisywanie przy użyciu xlwt
    '''spisik = xlwt.Workbook(encoding="utf-8")
    sheet = spisik.add_sheet("Spis konkurencji")
    sheet.write(0, 0, "Produkt")
    sheet.write(0, 1, "Cena")

    i = 1

    for prod in lista_prod:
        i += 1
        sheet.write(i, 0, prod)

    j = 1

    for cena in lista_cen:
        j += 1
        sheet.write(j, 1, cena)

    spisik.save("Spisik.xls")'''

    # zapisywanie przy użyciu openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Produkt')
    ws.cell(row=1, column=2, value='Cena')

    i = 3

    for prod in lista_prod:
        ws.cell(row=i, column=1, value=prod)
        i += 1

    j = 3

    for cena in lista_cen:
        ws.cell(row=j, column=2, value=cena)
        j += 1

    wb.save("Spisik.xlsx")


def spiscomparison():

    wb1 = openpyxl.load_workbook(filename='Spis konkurencji.xlsx')

    wb2 = openpyxl.load_workbook(filename='Spisik.xlsx')

    ws = wb1.get_sheet_by_name('Sheet1')

    spisik = wb2.get_sheet_by_name('Sheet1')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1):
        for cell in row:
            if cell.value in compdict:
                for wiersz in spisik.iter_rows(min_row=1, max_row=spisik.max_row, min_col=1):
                    for komorka in wiersz:
                        if komorka.value == compdict[cell.value]:
                            cena = spisik.cell(row=komorka.row, column=2)
                            ws.cell(row=cell.row, column=2, value=cena.value)

    wb1.save('Spis konkurencji.xlsx')
    wb2.close()


def saveresults_dg3():
    # zapisanie wyników spisu DG3 do formatki excela
    pass


def spis():
    # główny blok spisu konkurencji

    czas_start = time.time()

    print("\nSpis konkurencji\n")
    printcat()

    while True:
        ans = input("Naciśnij ENTER aby spisać kategorię; Wpisz 'exit' aby zakończyć:  \n")
        if ans == 'exit':
            break
        else:
            k = kategoria()
            synchro = pricesynchro(k)
            strony = get_max_pages(k)
            saveresults_spis(strony, k, synchro)

    czas_end = time.time()
    czas_wykonania = czas_end - czas_start
    print("Czas wykonania: " + str(czas_wykonania))


def degjetrzy():
    # główny blok spisu DG3
    pass


if __name__ == "__main__":

    spis()

    # wykonanie pełnego spisu RTVEuroAGD, bez synchronizacji cen
    '''czas1 = time.time()

    fulldict = {}

    for d in [RTV_url_dict, AGD_url_dict, AGD_zabudowa_url_dict, AGD_male_url_dict, Komputery_url_dict,
    Gry_i_konsole_url_dict, Foto_i_kamery_url_dict, Telefony_i_GPS_url_dict]:
        fulldict.update(d)

    wb = openpyxl.Workbook()
    ws = wb.active

    wiersz1, wiersz2 = 1, 1

    for kat in fulldict:
        kategoria = fulldict[kat]
        synchro = pricesynchro(kategoria)
        strony = get_max_pages(kategoria)
        nazwy_produktow(strony, kategoria)
        ceny_produktow(strony, kategoria, synchro)
        lista_prod = [x.encode('utf-8') for x in prod_temp]
        lista_cen = [x.encode('utf-8') for x in ceny_temp]

        for prod in lista_prod:
            ws.cell(row=wiersz1, column=1, value=prod)
            wiersz1 += 1

        for cena in lista_cen:
            ws.cell(row=wiersz2, column=2, value=cena)
            wiersz2 += 1

        prod_temp = []
        ceny_temp = []

    wb.save('spistest.xlsx')

    czas2 = time.time()

    finaltime = czas2 - czas1

    print finaltime'''

    # kat = kategoria()
    # parametry_produktow(1, kat)
    # sy = param_synchro(kat)
    # print sy

    # lista_par = [x.encode('utf-8') for x in param_temp]
    # chunks = [lista_par[x:x + 5] for x in xrange(0, len(lista_par), 5)]

    # wb = openpyxl.load_workbook('DG3formatka.xlsx')
    # sheet = wb.get_sheet_by_name('Spis')
    # sheet['B50'] = 'txesty'
    # wb.save('DG3.xlsx')
