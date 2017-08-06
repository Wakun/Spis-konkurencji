# coding=utf-8
"""
csvtest.csv - baza danych
inputtest.xlsx - plik do którego chcemy zapisać wynik spisu
spistest.xlsx - spis ze strony
"""

import openpyxl
import numpy as np

lines = np.genfromtxt("csvtest.csv", delimiter=";", dtype=None)
compdict = dict()
for i in range(len(lines)):
    compdict[lines[i][0]] = lines[i][1]

wb1 = openpyxl.load_workbook(filename='inputtest.xlsx')

wb2 = openpyxl.load_workbook(filename='spistest.xlsx')

ws = wb1.get_sheet_by_name('Sheet1')

spis = wb2.get_sheet_by_name('Sheet1')

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1):
    for cell in row:
        if cell.value in compdict:
            for wiersz in spis.iter_rows(min_row=1, max_row=spis.max_row, min_col=1):
                for komorka in wiersz:
                    if komorka.value == compdict[cell.value]:
                        cena = spis.cell(row=komorka.row, column=2)
                        ws.cell(row=cell.row, column=2, value=cena.value)

wb1.save('inputtest.xlsx')
wb2.close()
